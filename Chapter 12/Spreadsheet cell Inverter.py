# Spreadsheet cell inverter.py
# Author: Tan Yi Xiang
# Source: Automate the Boring stuff with python Ch. 12 Project
import openpyxl
import re

excelFilePath = input('Input the path of the excel file(with .xlsx extension):\n')

path_regex = re.compile(r'(.*/)(.*)(\.xlsx)$')
path_split = path_regex.search(excelFilePath)
path = path_split.group(1)
name = path_split.group(2)

resultPath = path + name + '(inverted).xlsx'

# Obtain values of original file cell
workbookToInvert = openpyxl.load_workbook(excelFilePath)
sheet = workbookToInvert.active
allColValues = []

for cols in range(1, sheet.max_column + 1):
    colValues = []
    for rows in range(1, sheet.max_row + 1):
        colValues.append(sheet.cell(rows, cols).value)
    allColValues.append(colValues)

# Transpose the values in new excel file
invertedWorkbook = openpyxl.Workbook()
invertSheet = invertedWorkbook.active
col_no = len(allColValues)

for cols in range(0, col_no):
    for cellNo in range(0, len(allColValues[cols])):
        invertSheet.cell(cols + 1, cellNo + 1).value = allColValues[cols][cellNo]

invertedWorkbook.save(resultPath)
print('The file have been transposed under:' + str(resultPath))
