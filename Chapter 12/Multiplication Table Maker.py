# Multiplication Table Maker.py
# Author: Tan Yi Xiang
# Source: Automate the Boring stuff with python Ch. 12 Project
import openpyxl
from openpyxl.styles import Font
import os

"Create a multiplication table in Excel from 1 to N "

N = int(input('What is the upper bound of the N * N multiplication table:?\n'))

workbook = openpyxl.Workbook()
tableSheet = workbook.active
tableSheet.title = str(N) + ' multiplication table'

# Create label rows and columns
for labelNum in range(1, N + 1):
    tableSheet.cell(row=1, column=labelNum + 1).value = labelNum
    tableSheet.cell(row=labelNum + 1, column=1).value = labelNum
    tableSheet.cell(row=1, column=labelNum + 1).font = Font(bold=True)
    tableSheet.cell(row=labelNum + 1, column=1).font = Font(bold=True)

# Fill in cells according to multiplication table
for columns in range(2, N + 2):
    for rows in range(2, N + 2):
        firstValue = tableSheet.cell(row=rows, column=1).value
        secondValue = tableSheet.cell(row=1, column=columns).value
        tableSheet.cell(row=rows, column=columns).value = firstValue * secondValue
workbook.save('Result.xlsx')
print('Multiplication table completed. Saved under ' + str(os.getcwd()) + "\Result.xlsx")
